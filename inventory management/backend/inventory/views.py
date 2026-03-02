from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db import transaction
from django.utils import timezone
from django.db.models import F, Sum
from django.http import HttpResponse
from datetime import timedelta
import csv
import secrets

from . import models, serializers
from .permissions import RolePermission, NoDeletePermission, NotViewerPermission
from .key_views import RoomKeyViewSet, KeyAuditLogViewSet, KeyBorrowViewSet
from django.conf import settings


class UserViewSet(viewsets.ModelViewSet):
    queryset = models.User.objects.all()
    serializer_class = serializers.UserSerializer
    filterset_fields = ['role']
    search_fields = ['name', 'email']
    permission_classes = [RolePermission]


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = models.Category.objects.all()
    serializer_class = serializers.CategorySerializer
    search_fields = ['category_name']
    permission_classes = [RolePermission]


class ItemViewSet(viewsets.ModelViewSet):
    queryset = models.Item.objects.select_related('category', 'room').all()
    serializer_class = serializers.ItemSerializer
    filterset_fields = ['category', 'room']
    search_fields = ['item_name']
    permission_classes = [RolePermission, NoDeletePermission]

    def create(self, request, *args, **kwargs):
        """
        Create item - Staff creates pending items, Admin creates directly
        """
        user = request.user
        
        # If user is staff (not admin/manager), create as pending item instead
        if user.role == 'staff':
            # Create pending item instead
            pending_data = request.data.copy()
            # Normalize field names from item payload
            if 'category_id' in pending_data and 'category' not in pending_data:
                pending_data['category'] = pending_data.get('category_id')
            if 'room_id' in pending_data and 'room' not in pending_data:
                pending_data['room'] = pending_data.get('room_id')
            pending_data['requested_by'] = user.pk
            pending_data['status'] = 'pending'
            
            pending_serializer = serializers.PendingItemSerializer(data=pending_data)
            pending_serializer.is_valid(raise_exception=True)
            pending_item = pending_serializer.save(requested_by=user)
            
            # Create audit log
            models.AuditLog.objects.create(
                user=user,
                action=f"Submitted new item for approval: {pending_item.item_name}"
            )
            
            return Response(
                {
                    'message': 'Item submitted for admin approval',
                    'pending_item': pending_serializer.data,
                    'status': 'pending_approval'
                },
                status=status.HTTP_201_CREATED
            )
        
        # Admin/Manager can create items directly
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        """Handle room assignment when creating items (for admin/manager only)"""
        item = serializer.save()
        # Create audit log for direct creation
        models.AuditLog.objects.create(
            user=self.request.user,
            action=f"Created item directly: {item.item_name}"
        )
        
    @action(detail=False, methods=['get'])
    def roomwise(self, request):
        """Get inventory grouped by room/location with room properties"""
        items = models.Item.objects.select_related('category', 'room').all().order_by('room__room_name', 'item_name')

        # Group items by room (key by room_id or 'unassigned')
        rooms_data = {}
        for item in items:
            if item.room:
                key = item.room.room_id
                if key not in rooms_data:
                    rooms_data[key] = {
                        'room_id': item.room.room_id,
                        'room_name': item.room.room_name,
                        'room_type': item.room.room_type,
                        'location': item.room.location,
                        'room_key': item.room.room_key,
                        'items': [],
                        'total_quantity': 0,
                        'item_count': 0
                    }
            else:
                key = 'unassigned'
                if key not in rooms_data:
                    rooms_data[key] = {
                        'room_id': None,
                        'room_name': 'General Storage',
                        'room_type': 'storage',
                        'location': '',
                        'room_key': False,
                        'items': [],
                        'total_quantity': 0,
                        'item_count': 0
                    }

            item_data = {
                'item_id': item.item_id,
                'item_name': item.item_name,
                'category': item.category.category_name if item.category else 'N/A',
                'unit': item.unit,
                'quantity': item.quantity,
                'min_quantity': item.min_quantity,
                'is_low_stock': item.quantity <= item.min_quantity,
                'description': item.description
            }

            rooms_data[key]['items'].append(item_data)
            rooms_data[key]['total_quantity'] += item.quantity
            rooms_data[key]['item_count'] += 1

        # Convert to list format sorted by room_name
        result = sorted(rooms_data.values(), key=lambda r: (r['room_name'] or ''))
        return Response(result)


class RoomViewSet(viewsets.ModelViewSet):
    queryset = models.Room.objects.all()
    serializer_class = serializers.RoomSerializer
    search_fields = ['room_name', 'room_key', 'location']
    filterset_fields = ['room_type']
    permission_classes = [RolePermission]


class RequisitionViewSet(viewsets.ModelViewSet):
    queryset = models.Requisition.objects.select_related('user').prefetch_related('items').all()
    serializer_class = serializers.RequisitionSerializer
    filterset_fields = ['status']
    permission_classes = [NotViewerPermission, RolePermission]

    def perform_create(self, serializer):
        # Ensure the requisition is associated with the authenticated user
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        requisition = self.get_object()
        if requisition.status != 'pending':
            return Response({'detail': 'Only pending requisitions can be approved.'}, status=400)
        requisition.status = 'approved'
        requisition.save(update_fields=['status'])
        models.AuditLog.objects.create(user=request.user if request.user.is_authenticated else None, action=f"Approved requisition #{requisition.req_id}")
        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        requisition = self.get_object()
        if requisition.status not in ['pending', 'approved']:
            return Response({'detail': 'Only pending/approved requisitions can be rejected.'}, status=400)
        requisition.status = 'rejected'
        requisition.save(update_fields=['status'])
        models.AuditLog.objects.create(user=request.user if request.user.is_authenticated else None, action=f"Rejected requisition #{requisition.req_id}")
        return Response({'status': 'rejected'})

    @action(detail=True, methods=['post'])
    def issue(self, request, pk=None):
        requisition = self.get_object()
        if requisition.status != 'approved':
            return Response({'detail': 'Only approved requisitions can be issued.'}, status=400)
        with transaction.atomic():
            for req_item in requisition.items.select_related('item').all():
                item = req_item.item
                if item.quantity < req_item.quantity:
                    return Response({'detail': f'Insufficient stock for {item.item_name}'}, status=400)
                item.quantity = item.quantity - req_item.quantity
                item.save(update_fields=['quantity'])
                models.StockTransaction.objects.create(
                    item=item,
                    type='OUT',
                    quantity=req_item.quantity,
                    user=request.user if request.user.is_authenticated else None,
                    notes=f"Issued for requisition #{requisition.req_id}"
                )
            requisition.status = 'issued'
            requisition.expected_return_at = timezone.now() + timedelta(days=requisition.return_duration_days or 0)
            requisition.save(update_fields=['status', 'expected_return_at'])
            models.AuditLog.objects.create(user=request.user if request.user.is_authenticated else None, action=f"Issued requisition #{requisition.req_id}")
        return Response({'status': 'issued'})

    @action(detail=True, methods=['post'], url_path='return')
    def mark_return(self, request, pk=None):
        requisition = self.get_object()
        if requisition.status != 'issued':
            return Response({'detail': 'Only issued requisitions can be returned.'}, status=400)

        requisition.status = 'returned'
        requisition.returned_at = timezone.now()
        requisition.save(update_fields=['status', 'returned_at'])

        models.AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=f"Returned requisition #{requisition.req_id}"
        )

        return Response({'status': 'returned'})


class RequisitionItemViewSet(viewsets.ModelViewSet):
    queryset = models.RequisitionItem.objects.select_related('requisition', 'item').all()
    serializer_class = serializers.RequisitionItemSerializer
    permission_classes = [RolePermission]


class StockTransactionViewSet(viewsets.ModelViewSet):
    queryset = models.StockTransaction.objects.select_related('item', 'user').all()
    serializer_class = serializers.StockTransactionSerializer
    filterset_fields = ['type', 'item']
    permission_classes = [RolePermission]

    def perform_create(self, serializer):
        with transaction.atomic():
            trans = serializer.save(user=self.request.user if self.request.user.is_authenticated else None)
            item = trans.item
            room_id = self.request.data.get('room')
            
            if trans.type == 'IN':
                item.quantity = item.quantity + trans.quantity
            elif trans.type == 'OUT':
                if item.quantity < trans.quantity:
                    raise Exception('Insufficient stock')
                item.quantity = item.quantity - trans.quantity
            
            # Update item's room if provided
            if room_id:
                try:
                    room = models.Room.objects.get(room_id=room_id)
                    item.room = room
                except models.Room.DoesNotExist:
                    pass
            
            item.save(update_fields=['quantity', 'room'])
            
            # Create room item history record
            if room_id and trans.type in ['IN', 'OUT']:
                try:
                    room = models.Room.objects.get(room_id=room_id)
                    models.RoomItemHistory.objects.create(
                        item=item,
                        from_room=item.room if trans.type == 'OUT' else None,
                        to_room=room if trans.type == 'IN' else None,
                        user=self.request.user if self.request.user.is_authenticated else None,
                        remarks=f"Stock {trans.type}: {trans.quantity} units. {trans.notes or ''}"
                    )
                except models.Room.DoesNotExist:
                    pass
            
            models.AuditLog.objects.create(
                user=self.request.user if self.request.user.is_authenticated else None,
                action=f"Stock {trans.type}: {item.item_name} x{trans.quantity}"
            )


class PendingStockTransactionViewSet(viewsets.ModelViewSet):
    """ViewSet for managing pending stock transactions awaiting approval"""
    queryset = models.PendingStockTransaction.objects.select_related('item', 'room', 'requested_by', 'approved_by').all()
    serializer_class = serializers.PendingStockTransactionSerializer
    filterset_fields = ['status', 'type', 'item']
    search_fields = ['item__item_name', 'requested_by__name']
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Only admins can see all pending requests; others see their own"""
        if self.request.user.role in ['admin', 'manager', 'staff']:
            return models.PendingStockTransaction.objects.select_related('item', 'room', 'requested_by', 'approved_by').all()
        return models.PendingStockTransaction.objects.select_related('item', 'room', 'requested_by', 'approved_by').filter(requested_by=self.request.user)

    def create(self, request, *args, **kwargs):
        """Create a pending stock transaction"""
        data = request.data.copy()
        data['requested_by'] = request.user.pk
        data['status'] = 'pending'
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def approve(self, request, pk=None):
        """Approve a pending stock transaction (admin only)"""
        if request.user.role not in ['admin', 'manager', 'staff']:
            return Response(
                {'detail': 'Only admins can approve stock transactions'},
                status=status.HTTP_403_FORBIDDEN
            )

        pending = self.get_object()
        if pending.status != 'pending':
            return Response(
                {'detail': f'Cannot approve non-pending transaction (current: {pending.status})'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            # Create the actual stock transaction
            item = pending.item
            if pending.type == 'IN':
                item.quantity = item.quantity + pending.quantity
            elif pending.type == 'OUT':
                if item.quantity < pending.quantity:
                    return Response(
                        {'detail': 'Insufficient stock to approve OUT transaction'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                item.quantity = item.quantity - pending.quantity

            # Update room if provided
            if pending.room:
                item.room = pending.room

            item.save(update_fields=['quantity', 'room'])

            # Create stock transaction record
            stock_trans = models.StockTransaction.objects.create(
                item=item,
                type=pending.type,
                quantity=pending.quantity,
                user=request.user,
                notes=f"Approved from pending: {pending.notes or ''}"
            )

            # Create room history if applicable
            if pending.room and pending.type in ['IN', 'OUT']:
                models.RoomItemHistory.objects.create(
                    item=item,
                    from_room=item.room if pending.type == 'OUT' else None,
                    to_room=pending.room if pending.type == 'IN' else None,
                    user=request.user,
                    remarks=f"Stock {pending.type}: {pending.quantity} units"
                )

            # Mark pending as approved
            pending.status = 'approved'
            pending.approved_by = request.user
            pending.approved_at = timezone.now()
            pending.save()

            models.AuditLog.objects.create(
                user=request.user,
                action=f"Approved pending stock {pending.type}: {item.item_name} x{pending.quantity}"
            )

        serializer = self.get_serializer(pending)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def reject(self, request, pk=None):
        """Reject a pending stock transaction"""
        if request.user.role not in ['admin', 'manager', 'staff']:
            return Response(
                {'detail': 'Only admins can reject transactions'},
                status=status.HTTP_403_FORBIDDEN
            )

        pending = self.get_object()
        if pending.status != 'pending':
            return Response(
                {'detail': f'Cannot reject non-pending transaction (current: {pending.status})'},
                status=status.HTTP_400_BAD_REQUEST
            )

        rejection_reason = request.data.get('rejection_reason', 'No reason provided')
        pending.status = 'rejected'
        pending.rejection_reason = rejection_reason
        pending.approved_by = request.user
        pending.approved_at = timezone.now()
        pending.save()

        models.AuditLog.objects.create(
            user=request.user,
            action=f"Rejected pending stock {pending.type}: {pending.item.item_name}. Reason: {rejection_reason}"
        )

        serializer = self.get_serializer(pending)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def pending_approvals(self, request):
        """Get all pending transactions awaiting approval (admin only)"""
        if request.user.role != 'admin':
            return Response(
                {'detail': 'Only admins can view pending approvals'},
                status=status.HTTP_403_FORBIDDEN
            )

        pending = models.PendingStockTransaction.objects.filter(status='pending').select_related('item', 'room', 'requested_by').order_by('requested_at')
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)


class PendingItemViewSet(viewsets.ModelViewSet):
    """ViewSet for managing pending item creation awaiting approval"""
    queryset = models.PendingItem.objects.select_related('category', 'room', 'requested_by', 'approved_by').all()
    serializer_class = serializers.PendingItemSerializer
    filterset_fields = ['status', 'category']
    search_fields = ['item_name', 'requested_by__name']
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Admins see all pending items; staff see their own"""
        if self.request.user.role in ['admin', 'manager']:
            return models.PendingItem.objects.select_related('category', 'room', 'requested_by', 'approved_by').all()
        return models.PendingItem.objects.select_related('category', 'room', 'requested_by', 'approved_by').filter(requested_by=self.request.user)

    def create(self, request, *args, **kwargs):
        """Create a pending item"""
        data = request.data.copy()
        data['requested_by'] = request.user.pk
        data['status'] = 'pending'
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def approve(self, request, pk=None):
        """Approve a pending item creation (admin/manager only)"""
        if request.user.role not in ['admin', 'manager']:
            return Response(
                {'detail': 'Only admins and managers can approve items'},
                status=status.HTTP_403_FORBIDDEN
            )

        pending = self.get_object()
        if pending.status != 'pending':
            return Response(
                {'detail': f'Cannot approve non-pending item (current: {pending.status})'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            normalized_name = (pending.item_name or '').strip()
            existing_qs = models.Item.objects.filter(item_name__iexact=normalized_name)
            if pending.room_id:
                existing_qs = existing_qs.filter(room=pending.room)
            else:
                existing_qs = existing_qs.filter(room__isnull=True)

            def normalize_unit(value):
                if value is None:
                    return None
                value = str(value).strip()
                return value.lower() if value else None

            pending_unit = normalize_unit(pending.unit)
            candidates = list(existing_qs)
            item = None

            for candidate in candidates:
                candidate_unit = normalize_unit(candidate.unit)
                category_ok = (
                    candidate.category_id == pending.category_id
                    if pending.category_id is not None
                    else candidate.category_id is None
                )
                unit_ok = (
                    candidate_unit == pending_unit
                    if pending_unit is not None
                    else candidate_unit is None
                )
                if category_ok and unit_ok:
                    item = candidate
                    break

            if item is None and candidates:
                item = candidates[0]
            if item:
                # Merge with existing inventory
                item.quantity = item.quantity + pending.quantity
                if pending.min_quantity and pending.min_quantity > item.min_quantity:
                    item.min_quantity = pending.min_quantity
                if pending.description and not item.description:
                    item.description = pending.description
                item.save(update_fields=['quantity', 'min_quantity', 'description'])
                audit_action = f"Approved and merged item: {item.item_name} (+{pending.quantity})"
            else:
                # Create the actual item
                item = models.Item.objects.create(
                    item_name=normalized_name,
                    category=pending.category,
                    room=pending.room,
                    unit=pending.unit,
                    quantity=pending.quantity,
                    min_quantity=pending.min_quantity,
                    description=pending.description
                )
                audit_action = f"Approved and created item: {item.item_name}"

            # Mark pending as approved
            pending.status = 'approved'
            pending.approved_by = request.user
            pending.approved_at = timezone.now()
            pending.created_item = item
            pending.save()

            models.AuditLog.objects.create(
                user=request.user,
                action=audit_action
            )

        serializer = self.get_serializer(pending)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated])
    def reject(self, request, pk=None):
        """Reject a pending item creation (admin/manager only)"""
        if request.user.role not in ['admin', 'manager']:
            return Response(
                {'detail': 'Only admins and managers can reject items'},
                status=status.HTTP_403_FORBIDDEN
            )

        pending = self.get_object()
        if pending.status != 'pending':
            return Response(
                {'detail': f'Cannot reject non-pending item (current: {pending.status})'},
                status=status.HTTP_400_BAD_REQUEST
            )

        rejection_reason = request.data.get('rejection_reason', 'No reason provided')
        pending.status = 'rejected'
        pending.rejection_reason = rejection_reason
        pending.approved_by = request.user
        pending.approved_at = timezone.now()
        pending.save()

        models.AuditLog.objects.create(
            user=request.user,
            action=f"Rejected pending item: {pending.item_name}. Reason: {rejection_reason}"
        )

        serializer = self.get_serializer(pending)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def pending_approvals(self, request):
        """Get all pending items awaiting approval (admin/manager only)"""
        if request.user.role not in ['admin', 'manager']:
            return Response(
                {'detail': 'Only admins and managers can view pending approvals'},
                status=status.HTTP_403_FORBIDDEN
            )

        pending = models.PendingItem.objects.filter(status='pending').select_related('category', 'room', 'requested_by').order_by('requested_at')
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)


class RoomItemHistoryViewSet(viewsets.ModelViewSet):
    queryset = models.RoomItemHistory.objects.select_related('item', 'from_room', 'to_room', 'user').all()
    serializer_class = serializers.RoomItemHistorySerializer
    permission_classes = [RolePermission]

    def perform_create(self, serializer):
        with transaction.atomic():
            history = serializer.save()
            # Update item's current room
            item = history.item
            item.room = history.to_room
            item.save(update_fields=['room'])
            # Audit
            models.AuditLog.objects.create(
                user=history.user,
                action=f"Moved '{item.item_name}' from {history.from_room} to {history.to_room}"
            )


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Audit Log ViewSet - Read-only access to audit logs"""
    queryset = models.AuditLog.objects.select_related('user').all()
    serializer_class = serializers.AuditLogSerializer
    permission_classes = [NotViewerPermission, RolePermission]
    filterset_fields = ['user']
    search_fields = ['action']


# Reports
from rest_framework.views import APIView

class DashboardStatsView(APIView):
    def get(self, request):
        # Base metrics
        total_items = models.Item.objects.count()
        low_stock_items = models.Item.objects.filter(quantity__lte=F('min_quantity')).count()
        available_items = models.Item.objects.filter(quantity__gt=0).count()

        # Month-to-date stock IN quantity
        now = timezone.now()
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        stock_in_month = models.StockTransaction.objects.filter(
            type='IN', timestamp__gte=start_of_month
        ).aggregate(total=Sum('quantity'))['total'] or 0

        data = {
            'total_items': total_items,
            'low_stock_items': low_stock_items,
            'available_items': available_items,
            'stock_in_month': stock_in_month,
            'pending_requisitions': models.Requisition.objects.filter(status='pending').count(),
        }
        return Response(data)


class RoomOverviewView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        now = timezone.now()
        since = now - timedelta(days=7)

        total_rooms = models.Room.objects.count()
        unassigned_items = models.Item.objects.filter(room__isnull=True).count()

        moves_qs = models.RoomItemHistory.objects.select_related(
            'item', 'from_room', 'to_room', 'user'
        ).filter(moved_at__gte=since).order_by('-moved_at')

        recent_moves_7d = moves_qs.count()
        recent_moves = [
            {
                'item_name': move.item.item_name,
                'from_room_name': move.from_room.room_name if move.from_room else None,
                'to_room_name': move.to_room.room_name if move.to_room else None,
                'user_name': move.user.name if move.user else 'System',
                'moved_at': move.moved_at,
            }
            for move in moves_qs[:10]
        ]

        return Response({
            'total_rooms': total_rooms,
            'unassigned_items': unassigned_items,
            'recent_moves_7d': recent_moves_7d,
            'recent_moves': recent_moves,
        })


class RoomwiseActivityView(APIView):
    """API endpoint for roomwise inventory page activity data"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        now = timezone.now()
        since_7d = now - timedelta(days=7)
        
        print(f"🔍 Activity API called at {now}")
        print(f"   Looking for data since: {since_7d}")

        # Recent stock transactions
        transactions_qs = models.StockTransaction.objects.select_related(
            'item', 'user'
        ).filter(timestamp__gte=since_7d).order_by('-timestamp')

        recent_transactions = [
            {
                'type': txn.type,
                'item_name': txn.item.item_name,
                'quantity': txn.quantity,
                'user_name': txn.user.name if txn.user else 'System',
                'timestamp': txn.timestamp,
                'notes': txn.notes or '',
            }
            for txn in transactions_qs[:10]
        ]

        # Recent room moves
        moves_qs = models.RoomItemHistory.objects.select_related(
            'item', 'from_room', 'to_room', 'user'
        ).filter(moved_at__gte=since_7d).order_by('-moved_at')

        recent_moves = [
            {
                'item_name': move.item.item_name,
                'from_room_name': move.from_room.room_name if move.from_room else 'Unassigned',
                'to_room_name': move.to_room.room_name if move.to_room else 'Unassigned',
                'user_name': move.user.name if move.user else 'System',
                'moved_at': move.moved_at,
                'remarks': move.remarks or '',
            }
            for move in moves_qs[:10]
        ]

        # Recent key borrows (requested, approved, borrowed, returned)
        key_borrows_qs = models.KeyBorrow.objects.select_related(
            'key', 'borrower', 'approver'
        ).filter(requested_at__gte=since_7d).order_by('-requested_at')
        
        print(f"   Key borrows found: {key_borrows_qs.count()}")

        recent_key_borrows = [
            {
                'key_number': borrow.key.key_number,
                'room_name': borrow.key.room_name,
                'borrower_name': borrow.borrower.name if borrow.borrower else 'Unknown',
                'borrower_email': borrow.borrower.email if borrow.borrower else None,
                'borrower_phone': borrow.borrower.phone_number if borrow.borrower else None,
                'borrower_department': borrow.borrower.department if borrow.borrower else None,
                'approver_name': borrow.approver.name if borrow.approver else None,
                'status': borrow.status,
                'requested_at': borrow.requested_at,
                'approved_at': borrow.approved_at,
                'borrowed_at': borrow.borrowed_at,
                'returned_at': borrow.returned_at,
                'purpose': borrow.purpose or '',
                'expected_return_at': borrow.expected_return_at,
            }
            for borrow in key_borrows_qs[:10]
        ]
        
        print(f"   Returning {len(recent_transactions)} transactions, {len(recent_moves)} moves, {len(recent_key_borrows)} key borrows")

        return Response({
            'recent_transactions': recent_transactions,
            'recent_moves': recent_moves,
            'recent_key_borrows': recent_key_borrows,
        })


class ExportCSVView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Optional timeframe filter (default: 90 days)
        try:
            days = int(request.query_params.get('days', 90))
        except (TypeError, ValueError):
            days = 90
        since = timezone.now() - timedelta(days=days)

        response = HttpResponse(content_type='text/csv')
        timestamp = timezone.now().strftime('%Y-%m-%d')
        response['Content-Disposition'] = f'attachment; filename="CSE_Inventory_Report_{timestamp}.csv"'
        writer = csv.writer(response)

        # Header
        writer.writerow(['CSE Inventory Management System - Report'])
        writer.writerow([f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        writer.writerow([f'Period: last {days} days'])
        writer.writerow([])

        # Low Stock Alerts
        writer.writerow(['LOW STOCK ALERTS'])
        writer.writerow(['Item Name', 'Current Stock', 'Minimum Stock', 'Shortage', 'Category'])
        low_items = models.Item.objects.select_related('category').filter(
            quantity__lte=models.models.F('min_quantity')
        ).order_by('item_name')
        for item in low_items:
            shortage = item.min_quantity - item.quantity
            cat = item.category.category_name if item.category else ''
            writer.writerow([item.item_name, item.quantity, item.min_quantity, shortage, cat])
        writer.writerow([])

        # Items by Category
        writer.writerow(['ITEMS BY CATEGORY'])
        writer.writerow(['Category', 'Count', 'Percentage'])
        total_items = models.Item.objects.count()
        from django.db.models import Count
        cat_rows = models.Item.objects.values('category__category_name').annotate(count=Count('item_id')).order_by('-count')
        for row in cat_rows:
            cat = row['category__category_name'] or 'Uncategorized'
            count = row['count']
            percent = f"{round((count / total_items) * 100, 1) if total_items else 0}%"
            writer.writerow([cat, count, percent])
        writer.writerow([])

        # All Items
        writer.writerow(['ALL ITEMS'])
        writer.writerow(['Item Name', 'Category', 'Quantity', 'Unit', 'Min Quantity', 'Status', 'Updated At'])
        all_items = models.Item.objects.select_related('category').all().order_by('item_name')
        for item in all_items:
            status_label = 'Low Stock' if item.quantity <= item.min_quantity else 'Available'
            cat = item.category.category_name if item.category else 'N/A'
            updated = getattr(item, 'updated_at', None)
            updated_str = updated.strftime('%Y-%m-%d %H:%M:%S') if updated else ''
            writer.writerow([item.item_name, cat, item.quantity, item.unit, item.min_quantity, status_label, updated_str])
        writer.writerow([])

        # Recent Stock Transactions
        writer.writerow(['STOCK TRANSACTIONS (Recent)'])
        writer.writerow(['Transaction ID', 'Item', 'Type', 'Quantity', 'User', 'Timestamp', 'Notes'])
        trans_qs = models.StockTransaction.objects.select_related('item', 'user').filter(timestamp__gte=since).order_by('-timestamp')[:500]
        for tr in trans_qs:
            user_name = tr.user.name if tr.user else ''
            writer.writerow([tr.transaction_id, tr.item.item_name, tr.type, tr.quantity, user_name, tr.timestamp.strftime('%Y-%m-%d %H:%M:%S'), tr.notes or ''])
        writer.writerow([])

        # Purchases
        writer.writerow(['PURCHASES (Recent)'])
        writer.writerow(['Purchase ID', 'Supplier', 'Item', 'Quantity', 'Unit Price', 'Total Cost', 'Purchase Date'])
        purchases_qs = models.Purchase.objects.select_related('supplier', 'item').filter(purchase_date__gte=since.date()).order_by('-purchase_date')[:500]
        for p in purchases_qs:
            writer.writerow([p.purchase_id, p.supplier.supplier_name, p.item.item_name, p.quantity, float(p.unit_price), float(p.total_cost), p.purchase_date.strftime('%Y-%m-%d')])
        writer.writerow([])

        # Requisitions
        writer.writerow(['REQUISITIONS (Recent)'])
        writer.writerow(['Req ID', 'User', 'Status', 'Created At', 'Purpose'])
        req_qs = models.Requisition.objects.select_related('user').filter(created_at__gte=since).order_by('-created_at')[:500]
        for r in req_qs:
            writer.writerow([r.req_id, r.user.name, r.status, r.created_at.strftime('%Y-%m-%d %H:%M:%S'), (r.purpose or '').replace('\n', ' ')])
        writer.writerow([])

        writer.writerow(['REQUISITION ITEMS (Recent)'])
        writer.writerow(['Req ID', 'Item', 'Quantity'])
        req_items_qs = models.RequisitionItem.objects.select_related('requisition', 'item').filter(requisition__created_at__gte=since).order_by('-req_item_id')[:1000]
        for ri in req_items_qs:
            writer.writerow([ri.requisition.req_id, ri.item.item_name, ri.quantity])

        return response


class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Optional timeframe filter (default: 90 days)
        try:
            days = int(request.query_params.get('days', 90))
        except (TypeError, ValueError):
            days = 90
        since = timezone.now() - timedelta(days=days)

        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws_summary['A1'] = 'CSE Inventory Management System - Report'
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary['A3'] = f'Period: last {days} days'
        
        # Sheet 2: Low Stock
        ws_low = wb.create_sheet("Low Stock")
        ws_low.append(['Item Name', 'Current Stock', 'Minimum Stock', 'Shortage', 'Category'])
        for cell in ws_low[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        low_items = models.Item.objects.select_related('category').filter(
            quantity__lte=F('min_quantity')
        ).order_by('item_name')
        for item in low_items:
            shortage = item.min_quantity - item.quantity
            cat = item.category.category_name if item.category else ''
            ws_low.append([item.item_name, item.quantity, item.min_quantity, shortage, cat])
        
        # Sheet 3: All Items
        ws_items = wb.create_sheet("All Items")
        ws_items.append(['Item Name', 'Category', 'Quantity', 'Unit', 'Min Quantity', 'Status', 'Updated At'])
        for cell in ws_items[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        all_items = models.Item.objects.select_related('category').all().order_by('item_name')
        for item in all_items:
            status_label = 'Low Stock' if item.quantity <= item.min_quantity else 'Available'
            cat = item.category.category_name if item.category else 'N/A'
            updated = getattr(item, 'updated_at', None)
            updated_str = updated.strftime('%Y-%m-%d %H:%M:%S') if updated else ''
            ws_items.append([item.item_name, cat, item.quantity, item.unit, item.min_quantity, status_label, updated_str])
        
        # Sheet 4: Transactions
        ws_trans = wb.create_sheet("Transactions")
        ws_trans.append(['Transaction ID', 'Item', 'Type', 'Quantity', 'User', 'Timestamp', 'Notes'])
        for cell in ws_trans[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        trans_qs = models.StockTransaction.objects.select_related('item', 'user').filter(
            timestamp__gte=since
        ).order_by('-timestamp')[:500]
        for tr in trans_qs:
            user_name = tr.user.name if tr.user else ''
            ws_trans.append([
                tr.transaction_id, tr.item.item_name, tr.type, tr.quantity,
                user_name, tr.timestamp.strftime('%Y-%m-%d %H:%M:%S'), tr.notes or ''
            ])
        
        # Sheet 5: Requisitions
        ws_req = wb.create_sheet("Requisitions")
        ws_req.append(['Req ID', 'User', 'Status', 'Created At', 'Purpose'])
        for cell in ws_req[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        req_qs = models.Requisition.objects.select_related('user').filter(
            created_at__gte=since
        ).order_by('-created_at')[:500]
        for r in req_qs:
            ws_req.append([
                r.req_id, r.user.name, r.status,
                r.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                (r.purpose or '').replace('\n', ' ')
            ])
        
        # Auto-size columns
        for ws in [ws_summary, ws_low, ws_items, ws_trans, ws_req]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        timestamp = timezone.now().strftime('%Y-%m-%d')
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="CSE_Inventory_Report_{timestamp}.xlsx"'
        
        return response


# Password Reset

class PasswordResetRequestView(generics.GenericAPIView):
    """Request password reset - generates token"""
    serializer_class = serializers.PasswordResetRequestSerializer
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        email = serializer.validated_data['email']
        user = models.User.objects.get(email=email)
        
        # Generate token
        token = secrets.token_urlsafe(32)
        expires_at = timezone.now() + timedelta(hours=1)
        
        # Invalidate old tokens
        models.PasswordResetToken.objects.filter(user=user, used=False).update(used=True)
        
        # Create new token
        reset_token = models.PasswordResetToken.objects.create(
            user=user,
            token=token,
            expires_at=expires_at
        )
        
        # Send email with reset link
        from django.core.mail import send_mail
        from django.conf import settings
        
        reset_url = f"{request.scheme}://{request.get_host()}/reset-password/?token={token}"
        
        subject = 'Password Reset Request - CSE Inventory System'
        message = f'''Hello {user.name},

You have requested to reset your password for the CSE Inventory Management System.

Click the link below to reset your password:
{reset_url}

This link will expire in 1 hour.

If you did not request this password reset, please ignore this email and your password will remain unchanged.

For security reasons, never share this link with anyone.

Best regards,
CSE Inventory System
'''
        
        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            return Response({
                'message': 'Password reset link has been sent to your email. Please check your inbox.',
            }, status=status.HTTP_200_OK)
        except Exception as e:
            # Log error but don't expose details to user
            print(f"Email sending failed: {str(e)}")
            return Response({
                'message': 'Password reset link has been sent to your email. Please check your inbox.',
            }, status=status.HTTP_200_OK)


class PasswordResetConfirmView(generics.GenericAPIView):
    """Confirm password reset with token"""
    serializer_class = serializers.PasswordResetConfirmSerializer
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        token_str = serializer.validated_data['token']
        new_password = serializer.validated_data['new_password']
        
        try:
            reset_token = models.PasswordResetToken.objects.get(token=token_str)
            
            if not reset_token.is_valid():
                return Response(
                    {'error': 'Token is invalid or has expired.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Reset password
            user = reset_token.user
            user.set_password(new_password)
            user.save()
            
            # Mark token as used
            reset_token.used = True
            reset_token.save()
            
            # Log action
            models.AuditLog.objects.create(
                user=user,
                action=f"Password reset completed for {user.email}"
            )
            
            return Response({
                'message': 'Password has been reset successfully. You can now login with your new password.'
            }, status=status.HTTP_200_OK)
            
        except models.PasswordResetToken.DoesNotExist:
            return Response(
                {'error': 'Invalid token.'},
                status=status.HTTP_400_BAD_REQUEST
            )


# Registration

class RegisterView(generics.CreateAPIView):
    queryset = models.User.objects.all()
    serializer_class = serializers.UserRegistrationSerializer
    permission_classes = [AllowAny]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            'user_id': user.user_id,
            'name': user.name,
            'email': user.email,
            'role': user.role,
            'message': 'Registration successful'
        }, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def current_user(request):
    if request.method == 'GET':
        serializer = serializers.UserSerializer(request.user)
        return Response(serializer.data)
    
    elif request.method == 'PATCH':
        serializer = serializers.UserSerializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
